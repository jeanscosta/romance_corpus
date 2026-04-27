"""
================================================================================
Boundary-Crossing Motion Event Encoding in L2 English
================================================================================
A corpus-based study of manner verb constructions and path verbs in learner
corpora across CEFR proficiency levels (A1–C1) and four Romance L1 backgrounds
(French, Italian, Portuguese, Spanish).

This script implements the full analysis pipeline:
  1. Corpus search for manner verb + into/out of constructions
  2. Corpus search for boundary-crossing path verbs (enter + exit)
  3. Automated semantic filtering via spatial Ground-noun lexicon
  4. Statistical analyses (Jonckheere-Terpstra, Spearman, Kruskal-Wallis)
  5. Figure generation (5 publication-quality PNG figures)
  6. Excel output (frequency tables)

Input:  POS-tagged corpus files in word_TAG format (Penn Treebank tagset)
        File naming: one file per L1 × CEFR level cell
        Configure paths in the FILES dictionary in Section 1.

Output: - Console: descriptive stats + inferential results
        - manner_verb_results.xlsx
        - Figure1_MV_into.png  through  Figure5_SM_CM.png
        - concordance_mv_into.txt, concordance_mv_outof.txt,
          concordance_pv.txt (for manual verification)

Requirements:
    pip install numpy scipy matplotlib openpyxl

Authors:  [Your name(s)]
Version:  2.0
License:  MIT
DOI:      [To be assigned]
================================================================================
"""

import re
import os
import numpy as np
from collections import defaultdict, Counter
from scipy import stats
from scipy.special import gammaln
from scipy.optimize import minimize
import warnings
warnings.filterwarnings('ignore')

# ╔══════════════════════════════════════════════════════════════════════════════
# 1. CONFIGURATION — edit this section to match your file structure
# ╚══════════════════════════════════════════════════════════════════════════════

# Update these paths to point to your POS-tagged corpus files.
# Keys: (language, CEFR_level) tuples.  Values: file paths.
FILES = {
    ('french',     'A1'): 'ST_french_A1_all.txt',
    ('french',     'A2'): 'ST_french_A2_all.txt',
    ('french',     'B1'): 'ST_french_B1_all.txt',
    ('french',     'B2'): 'ST_french_B2_all.txt',
    ('french',     'C1'): 'ST_french_C1_all.txt',
    ('italian',    'A1'): 'ST_italian_A1_all.txt',
    ('italian',    'A2'): 'ST_italian_A2_all.txt',
    ('italian',    'B1'): 'ST_italian_B1_all.txt',
    ('italian',    'B2'): 'ST_italian_B2_all.txt',
    ('italian',    'C1'): 'ST_italian_C1_all.txt',
    ('portuguese', 'A1'): 'ST_portuguese_A1_all.txt',
    ('portuguese', 'A2'): 'ST_portuguese_A2_all.txt',
    ('portuguese', 'B1'): 'ST_portuguese_B1_all.txt',
    ('portuguese', 'B2'): 'ST_portuguese_B2_all.txt',
    ('portuguese', 'C1'): 'ST_portuguese_C1_all.txt',
    ('spanish',    'A1'): 'ST_spanish_A1_all.txt',
    ('spanish',    'A2'): 'ST_spanish_A2_all.txt',
    ('spanish',    'B1'): 'ST_spanish_B1_all.txt',
    ('spanish',    'B2'): 'ST_spanish_B2_all.txt',
    ('spanish',    'C1'): 'ST_spanish_C1_all.txt',
}

LANGUAGES = ['french', 'italian', 'portuguese', 'spanish']
LEVELS    = ['A1', 'A2', 'B1', 'B2', 'C1']
LEVEL_MAP = {'A1': 1, 'A2': 2, 'B1': 3, 'B2': 4, 'C1': 5}

# Look-ahead window (tokens) for preposition search
WINDOW = 4

# ╔══════════════════════════════════════════════════════════════════════════════
# 2. MANNER VERB LISTS
#    Grounded in Talmy (1985, 2000) and Slobin (1996).
#    Add or remove lemmas as needed; update LEMMA_MAP with inflected forms.
# ╚══════════════════════════════════════════════════════════════════════════════

SELF_MOTION_VERBS = {
    # Base forms
    'walk', 'run', 'jump', 'rush', 'sneak', 'crawl', 'climb', 'swim', 'fly',
    'march', 'slide', 'slip', 'stumble', 'tiptoe', 'dash', 'race', 'sprint',
    'stroll', 'wander', 'skip', 'hop', 'limp', 'dive', 'burst', 'crash',
    # Inflected forms
    'walked','walking','walks', 'ran','running','runs',
    'jumped','jumping','jumps', 'rushed','rushing','rushes',
    'sneaked','sneaking','sneaks','snuck',
    'crawled','crawling','crawls', 'climbed','climbing','climbs',
    'swam','swimming','swims', 'flew','flying','flies',
    'marched','marching','marches', 'slid','sliding','slides',
    'slipped','slipping','slips', 'stumbled','stumbling',
    'tiptoed','tiptoeing', 'dashed','dashing','dashes',
    'raced','racing','races', 'sprinted','sprinting','sprints',
    'strolled','strolling','strolls', 'wandered','wandering','wanders',
    'skipped','skipping','skips', 'hopped','hopping','hops',
    'limped','limping','limps', 'dived','dove','diving','dives',
    'bursting','bursts', 'crashed','crashing','crashes',
}

CAUSED_MOTION_VERBS = {
    # Base forms
    'throw', 'kick', 'push', 'pull', 'hit', 'toss', 'roll', 'drop', 'pour',
    'drag', 'chase', 'bounce', 'shoot',
    # Inflected forms
    'threw','throwing','throws','thrown',
    'kicked','kicking','kicks', 'pushed','pushing','pushes',
    'pulled','pulling','pulls', 'hitting','hits',
    'tossed','tossing','tosses', 'rolled','rolling','rolls',
    'dropped','dropping','drops', 'poured','pouring','pours',
    'dragged','dragging','drags', 'chased','chasing','chases',
    'bounced','bouncing','bounces', 'shot','shooting','shoots',
}

MANNER_VERBS = SELF_MOTION_VERBS | CAUSED_MOTION_VERBS

# Maps each inflected form to its base lemma
LEMMA_MAP = {
    'walked':'walk','walking':'walk','walks':'walk',
    'ran':'run','running':'run','runs':'run',
    'jumped':'jump','jumping':'jump','jumps':'jump',
    'rushed':'rush','rushing':'rush','rushes':'rush',
    'sneaked':'sneak','sneaking':'sneak','sneaks':'sneak','snuck':'sneak',
    'crawled':'crawl','crawling':'crawl','crawls':'crawl',
    'climbed':'climb','climbing':'climb','climbs':'climb',
    'swam':'swim','swimming':'swim','swims':'swim',
    'flew':'fly','flying':'fly','flies':'fly',
    'marched':'march','marching':'march','marches':'march',
    'slid':'slide','sliding':'slide','slides':'slide',
    'slipped':'slip','slipping':'slip','slips':'slip',
    'stumbled':'stumble','stumbling':'stumble',
    'tiptoed':'tiptoe','tiptoeing':'tiptoe',
    'dashed':'dash','dashing':'dash','dashes':'dash',
    'raced':'race','racing':'race','races':'race',
    'sprinted':'sprint','sprinting':'sprint','sprints':'sprint',
    'strolled':'stroll','strolling':'stroll','strolls':'stroll',
    'wandered':'wander','wandering':'wander','wanders':'wander',
    'skipped':'skip','skipping':'skip','skips':'skip',
    'hopped':'hop','hopping':'hop','hops':'hop',
    'limped':'limp','limping':'limp','limps':'limp',
    'dived':'dive','dove':'dive','diving':'dive','dives':'dive',
    'bursting':'burst','bursts':'burst',
    'crashed':'crash','crashing':'crash','crashes':'crash',
    'threw':'throw','throwing':'throw','throws':'throw','thrown':'throw',
    'kicked':'kick','kicking':'kick','kicks':'kick',
    'pushed':'push','pushing':'push','pushes':'push',
    'pulled':'pull','pulling':'pull','pulls':'pull',
    'hitting':'hit','hits':'hit',
    'tossed':'toss','tossing':'toss','tosses':'toss',
    'rolled':'roll','rolling':'roll','rolls':'roll',
    'dropped':'drop','dropping':'drop','drops':'drop',
    'poured':'pour','pouring':'pour','pours':'pour',
    'dragged':'drag','dragging':'drag','drags':'drag',
    'chased':'chase','chasing':'chase','chases':'chase',
    'bounced':'bounce','bouncing':'bounce','bounces':'bounce',
    'shot':'shoot','shooting':'shoot','shoots':'shoot',
}

def get_lemma(verb):
    """Return base lemma for an inflected verb form."""
    return LEMMA_MAP.get(verb, verb)

def get_motion_type(lemma):
    """Classify a lemma as self-motion or caused-motion."""
    base = LEMMA_MAP.get(lemma, lemma)
    if base in SELF_MOTION_VERBS:   return 'self-motion'
    if base in CAUSED_MOTION_VERBS: return 'caused-motion'
    return 'unknown'

# ╔══════════════════════════════════════════════════════════════════════════════
# 3. PATH VERB FORMS
#    Only enter and exit are included as boundary-crossing path verbs.
#    leave, arrive, depart excluded (see paper Methods §1.3 for rationale).
# ╚══════════════════════════════════════════════════════════════════════════════

PATH_FORMS = {
    'enter', 'enters', 'entered', 'entering',
    'exit',  'exits',  'exited',  'exiting',
}
PATH_TO_LEMMA = {f: ('enter' if 'enter' in f else 'exit') for f in PATH_FORMS}

# ╔══════════════════════════════════════════════════════════════════════════════
# 4. GROUND NOUN LEXICONS FOR SEMANTIC FILTERING
#    These were built iteratively through automated classification +
#    manual review (see paper Methods §1.5 and supplementary materials).
#    To extend to other languages or corpora, add nouns to SPATIAL_GROUNDS
#    and to the exclusion sets as appropriate.
# ╚══════════════════════════════════════════════════════════════════════════════

SPATIAL_GROUNDS = {
    # ── Enclosed structures / buildings ───────────────────────────────────────
    'room', 'house', 'building', 'office', 'store', 'shop', 'kitchen',
    'bedroom', 'bathroom', 'hallway', 'corridor', 'garage', 'basement',
    'attic', 'closet', 'elevator', 'lift', 'classroom', 'library',
    'hospital', 'church', 'bar', 'restaurant', 'café', 'cafe', 'stadium',
    'arena', 'theater', 'theatre', 'cinema', 'gym', 'pool', 'jail',
    'prison', 'cell', 'tent', 'cave', 'tunnel', 'lobby', 'entrance',
    'doorway', 'alley', 'apartment', 'flat', 'cottage', 'cabin', 'mansion',
    'palace', 'castle', 'warehouse', 'factory', 'laboratory', 'museum',
    'gallery', 'hotel', 'inn', 'pub', 'nightclub', 'club', 'booth',
    'shelter', 'bunker', 'vault', 'chamber', 'hall', 'auditorium',
    'gymnasium', 'depot', 'station', 'terminal', 'hangar', 'barn', 'shed',
    'kennel', 'cage',
    # ── Typo/variant forms of the above ──────────────────────────────────────
    'home', 'homes', 'houses', 'living', 'porch', 'door', 'doors', 'bunk',
    'apartament', 'rooms', 'roon', 'thehall',
    # ── Outdoor bounded spaces ────────────────────────────────────────────────
    'garden', 'yard', 'court', 'courtyard', 'field', 'pitch', 'playground',
    'park', 'enclosure', 'compound', 'camp', 'site', 'area', 'zone',
    'ring', 'rink', 'track', 'lane', 'pen', 'paddock', 'bounds',
    # ── Educational / institutional buildings ────────────────────────────────
    'university', 'school', 'college', 'campus', 'faculty', 'lab',
    'laboratory', 'ufs', 'esci',
    # ── Natural spaces and environments ──────────────────────────────────────
    'water', 'river', 'lake', 'sea', 'ocean', 'pond', 'stream', 'creek',
    'canal', 'bay', 'harbor', 'harbour', 'beach', 'jungle', 'forest',
    'wood', 'woods', 'bush', 'clearing', 'valley', 'canyon', 'wilderness',
    'mountains', 'mountain', 'iceberg', 'coral', 'watr',
    # ── Vehicles and containers ───────────────────────────────────────────────
    'car', 'bus', 'train', 'plane', 'boat', 'ship', 'truck', 'van', 'cab',
    'taxi', 'vehicle', 'container', 'box', 'bin', 'basket', 'bag', 'bucket',
    'bowl', 'tank', 'barrel', 'bottle', 'jar', 'cup', 'pot', 'crate',
    'drawer', 'trunk', 'bottles', 'baskets', 'cesta', 'ark', 'airplane',
    'aircraft',
    # ── Urban spaces / infrastructure ────────────────────────────────────────
    'street', 'road', 'square', 'market', 'mall', 'village', 'town', 'city',
    'neighborhood', 'neighbourhood', 'district', 'quarter', 'block',
    'intersection', 'bridge', 'path', 'trail', 'passage', 'streed',
    'lamppost', 'lamp', 'bank', 'corner', 'touwn',
    # ── Sports / game bounds ──────────────────────────────────────────────────
    'goal', 'net', 'hole', 'gol', 'arc', 'circles', 'circle', 'court',
    # ── Geographic territories (boundary-crossing sense) ─────────────────────
    'country', 'countries', 'iran', 'usa', 'u.s.', 'us', 'eua', 'u.s',
    # ── Open/generic spaces ───────────────────────────────────────────────────
    'air', 'space', 'region', 'territory', 'land', 'ground', 'soil', 'sand',
    'mud', 'snow', 'ice', 'grass', 'dirt', 'floor', 'ceiling', 'wall',
    'place', 'world', 'narnia',
    # ── Other verified spatial nouns (from manual review) ────────────────────
    'exit', 'studio', 'zoo', 'department', 'departamento', 'window',
    'windows', 'bed', 'airport', 'shopping', 'movies', 'dining', 'sports',
    'medicine', 'cages', 'buildings', 'lhe', 'em', 'forum', 'aparment',
    'ball', 'thief', 'name', 'word', 'error', 'redaction', 'translator',
    'cash', 'facudade', 'hospitality', 'environnement', 'animals',
    'elephantes', 'croupiers', 'provideurs', 'pay', 'stardom', 'army',
    'experiences',
}

# Nouns to always exclude (body parts, persons, abstractions)
EXCLUDE_NOUNS = {
    # Body parts
    'stomach', 'belly', 'chest', 'back', 'face', 'mouth', 'eye', 'eyes',
    'ear', 'ears', 'nose', 'throat', 'neck', 'shoulder', 'shoulders', 'arm',
    'arms', 'hand', 'hands', 'leg', 'legs', 'foot', 'feet', 'knee', 'knees',
    'hip', 'hips', 'head', 'body', 'heart', 'skin', 'side', 'rib', 'ribs',
    'gut', 'guts', 'lap', 'wrist', 'ankle', 'spine', 'bone', 'bones',
    'finger', 'fingers', 'toe', 'toes', 'cheek', 'jaw', 'forehead', 'temple',
    'skull', 'brain', 'lung', 'lungs', 'liver', 'kidney',
    # Person / animate nouns
    'meg', 'person', 'people', 'man', 'woman', 'girl', 'boy', 'child',
    'children', 'someone', 'anyone', 'everybody', 'opponent', 'friend',
    'player', 'referee', 'coach', 'teacher', 'student', 'worker', 'boss',
    'colleague', 'mother', 'father', 'parent', 'sister', 'brother', 'baby',
    'kid', 'adult', 'human', 'enemy', 'victim', 'attacker', 'suspect',
    'officer', 'soldier', 'guard',
    # Abstract / idiomatic (MV constructions)
    'play', 'tears', 'laughter', 'song', 'silence', 'rage', 'fury', 'anger',
    'panic', 'despair', 'grief', 'joy', 'ecstasy', 'shock', 'surprise',
    'confusion', 'trouble', 'problems', 'problem', 'difficulty',
    'difficulties', 'danger', 'chaos', 'turmoil', 'crisis', 'conflict',
    'debate', 'discussion', 'argument', 'love', 'hate', 'fear', 'doubt',
    'hope', 'faith', 'belief', 'thought', 'idea', 'question', 'answer',
    'detail', 'details', 'depth', 'focus', 'account', 'consideration',
    'context', 'contact', 'competition', 'existence', 'reality', 'history',
    'tradition', 'culture', 'politics', 'business', 'industry', 'trade',
    'category', 'type', 'group', 'class', 'level', 'stage', 'phase', 'mode',
    'style', 'form', 'shape', 'action', 'effect', 'operation', 'production',
    'service', 'force', 'line', 'row', 'file', 'order', 'nature', 'density',
    'energy', 'time', 'money', 'fuels', 'fuel', 'drinking', 'january',
    'celebrities', 'groups', 'camel', 'ad', 'quest', 'persone', 'icy',
    'm.', 'tires',
    # Abstract / idiomatic (path verbs)
    'contact', 'agreement', 'deal', 'contract', 'workforce', 'labor',
    'labour', 'job', 'phd', 'master', 'study', 'studies', 'course',
    'program', 'programme', 'activities', 'data', 'right', 'things',
    'moment', 'arrival', 'day', 'party', 'multinacional', 'masters', 'power',
    'cook', 'drug', 'credentials', 'women', 'segment', 'horse', 'part',
    'bike', 'usd', 'market',
}

VERB_TAGS = {'VB', 'VBZ', 'VBP', 'VBD', 'VBG', 'VBN'}
DET_TAGS  = {'DT', 'PRP$', 'CD', 'JJ', 'JJR', 'JJS', 'POS', 'WDT'}


def ground_is_spatial(noun):
    """Return True if noun is a verified spatial ground."""
    if noun is None:
        return False
    n = noun.lower()
    return n in SPATIAL_GROUNDS and n not in EXCLUDE_NOUNS


def get_ground_noun(window_tagged, start_idx):
    """
    Extract the head noun from a window_tagged list starting at start_idx.
    Skips determiners and adjectives; returns the first noun found, or None.
    """
    for k in range(start_idx, min(start_idx + 5, len(window_tagged))):
        nw, nt = window_tagged[k]
        if nt in {'NN', 'NNS', 'NNP', 'NNPS'}:
            return nw
        elif nt not in DET_TAGS:
            return None
    return None


# ╔══════════════════════════════════════════════════════════════════════════════
# 5. CORPUS SEARCH
# ╚══════════════════════════════════════════════════════════════════════════════

def search_file(path):
    """
    Search a single POS-tagged corpus file for all three target constructions.

    Parameters
    ----------
    path : str  Path to a POS-tagged file (word_TAG format, space-separated).

    Returns
    -------
    dict with keys:
        wc              : int   word count (alphabetic tokens)
        mv_into         : list  [(lemma, ground_noun, context_str), ...]
        mv_outof        : list  [(lemma, ground_noun, context_str), ...]
        pv_spatial      : list  [(lemma, ground_noun, context_str), ...]
        pv_no_object    : int   count of enter/exit with no recoverable ground
    """
    with open(path, encoding='utf-8') as f:
        tokens = f.read().split()

    wc = sum(1 for t in tokens if re.match(r'^[A-Za-z]', t))
    mv_into   = []
    mv_outof  = []
    pv_spatial   = []
    pv_no_object = 0

    for i, token in enumerate(tokens):
        parts = token.rsplit('_', 1)
        if len(parts) != 2:
            continue
        word, tag = parts
        w = word.lower()

        # ── Manner verb constructions ─────────────────────────────────────────
        if tag in VERB_TAGS and w in MANNER_VERBS:
            lemma   = get_lemma(w)
            context = ' '.join(tokens[max(0, i - 2): i + WINDOW + 3])
            wt = [(t.rsplit('_', 1)[0].lower(), t.rsplit('_', 1)[1])
                  for t in tokens[i + 1: i + WINDOW + 2]
                  if len(t.rsplit('_', 1)) == 2]

            # MV + into
            for j, (nw, nt) in enumerate(wt):
                if nw == 'into' and nt == 'IN':
                    ground = get_ground_noun(wt, j + 1)
                    if ground_is_spatial(ground):
                        mv_into.append((lemma, ground, context))
                    break

            # MV + out of
            for j, (nw, nt) in enumerate(wt):
                if nw == 'out' and nt in ('IN', 'RP'):
                    if j + 1 < len(wt) and wt[j + 1][0] == 'of':
                        ground = get_ground_noun(wt, j + 2)
                        if ground_is_spatial(ground):
                            mv_outof.append((lemma, ground, context))
                    break

        # ── Path verbs ────────────────────────────────────────────────────────
        if tag in VERB_TAGS and w in PATH_FORMS:
            lemma   = PATH_TO_LEMMA[w]
            context = ' '.join(tokens[max(0, i - 2): i + 6])
            wt = [(t.rsplit('_', 1)[0].lower(), t.rsplit('_', 1)[1])
                  for t in tokens[i + 1: i + 6]
                  if len(t.rsplit('_', 1)) == 2]

            ground = get_ground_noun(wt, 0)
            if ground_is_spatial(ground):
                pv_spatial.append((lemma, ground, context))
            else:
                pv_no_object += 1

    return {
        'wc': wc,
        'mv_into':      mv_into,
        'mv_outof':     mv_outof,
        'pv_spatial':   pv_spatial,
        'pv_no_object': pv_no_object,
    }


def run_corpus_search():
    """Run search across all corpus files. Returns nested dict of results."""
    results = {}
    print(f"\n{'Cell':<25} {'Words':>10} {'MV+into':>9} {'MV+outof':>10} {'PV':>6}")
    print('-' * 65)

    for lang in LANGUAGES:
        for level in LEVELS:
            key  = (lang, level)
            path = FILES[key]
            if not os.path.exists(path):
                print(f"  WARNING: not found: {path}")
                continue
            r  = search_file(path)
            wc = r['wc']
            ni = len(r['mv_into'])
            no = len(r['mv_outof'])
            np_ = len(r['pv_spatial'])

            def norm(x):
                return round(x / wc * 10000, 4) if wc else 0

            results[key] = {
                'lang': lang, 'level': level,
                'level_num': LEVEL_MAP[level], 'wc': wc,
                'raw_mv_into':  ni,  'norm_mv_into':  norm(ni),
                'raw_mv_outof': no,  'norm_mv_outof': norm(no),
                'raw_mv_total': ni + no, 'norm_mv_total': norm(ni + no),
                'raw_pv':       np_, 'norm_pv':       norm(np_),
                'raw_pv_no_object': r['pv_no_object'],
                'mv_into_hits':  r['mv_into'],
                'mv_outof_hits': r['mv_outof'],
                'pv_hits':       r['pv_spatial'],
            }
            print(f"{lang}_{level:<18} {wc:>10,} "
                  f"{ni:>5}({norm(ni):.3f}) "
                  f"{no:>5}({norm(no):.3f}) "
                  f"{np_:>4}({norm(np_):.3f})")

    return results


# ╔══════════════════════════════════════════════════════════════════════════════
# 6. STATISTICAL ANALYSES
# ╚══════════════════════════════════════════════════════════════════════════════

def jonckheere_terpstra(groups):
    """
    Jonckheere-Terpstra trend test for ordered alternatives.
    One-tailed test of H1: group_1 ≤ group_2 ≤ ... ≤ group_k

    Parameters
    ----------
    groups : list of lists  Ordered from smallest to largest treatment level.

    Returns
    -------
    J : float   Test statistic
    z : float   Standardized statistic
    p : float   One-tailed p-value (upper tail = increasing direction)
    """
    J = 0
    for i in range(len(groups)):
        for j in range(i + 1, len(groups)):
            for x in groups[i]:
                for y in groups[j]:
                    if y > x:
                        J += 1
                    elif y == x:
                        J += 0.5

    n  = sum(len(g) for g in groups)
    ns = [len(g) for g in groups]
    E  = (n ** 2 - sum(ni ** 2 for ni in ns)) / 4
    V  = (n ** 2 * (2 * n + 3) - sum(ni ** 2 * (2 * ni + 3) for ni in ns)) / 72
    z  = (J - E) / np.sqrt(V)
    p  = 1 - stats.norm.cdf(z)
    return J, z, p


def partial_spearman(x, y, z):
    """
    Partial Spearman correlation between x and y, controlling for z.
    Uses the residualization method on ranked values.
    """
    rx = stats.rankdata(x)
    ry = stats.rankdata(y)
    rz = stats.rankdata(z)
    sx, ix, _, _, _ = stats.linregress(rz, rx)
    sy, iy, _, _, _ = stats.linregress(rz, ry)
    rx_r = rx - (sx * rz + ix)
    ry_r = ry - (sy * rz + iy)
    r, _ = stats.pearsonr(rx_r, ry_r)
    n = len(x)
    t = r * np.sqrt((n - 3) / (1 - r ** 2))
    p = 2 * stats.t.sf(abs(t), df=n - 3)
    return r, p


def run_statistics(results):
    """Run all statistical analyses and print results."""
    rows       = [results[(l, v)] for l in LANGUAGES for v in LEVELS]
    level_nums = np.array([d['level_num']   for d in rows])
    norm_into  = np.array([d['norm_mv_into'] for d in rows])
    norm_outof = np.array([d['norm_mv_outof']for d in rows])
    norm_pv    = np.array([d['norm_pv']      for d in rows])
    norm_mv    = np.array([d['norm_mv_total']for d in rows])

    sig = lambda p: '***' if p < .001 else '**' if p < .01 else '*' if p < .05 else 'n.s.'

    for label, norm in [('MV+into', norm_into),
                        ('MV+out of', norm_outof),
                        ('Path verbs (enter+exit)', norm_pv)]:
        print(f"\n{'='*65}")
        print(f"  {label}")
        print(f"{'='*65}")

        print("\n  Descriptive (mean normalized freq per level):")
        for lvl in LEVELS:
            vals = [d[('norm_mv_into' if 'into' in label
                        else 'norm_mv_outof' if 'out' in label
                        else 'norm_pv')]
                    for d in rows if d['level'] == lvl]
            print(f"    {lvl}: M={np.mean(vals):.4f}  SD={np.std(vals):.4f}  "
                  f"range=[{min(vals):.4f}, {max(vals):.4f}]")

        # J-T
        level_groups = [[d[('norm_mv_into' if 'into' in label
                              else 'norm_mv_outof' if 'out' in label
                              else 'norm_pv')]
                          for d in rows if d['level'] == lvl]
                         for lvl in LEVELS]
        J, z, p = jonckheere_terpstra(level_groups)
        print(f"\n  Jonckheere-Terpstra: J={J:.1f}, z={z:.3f}, "
              f"p={p:.4f}  {sig(p)}")

        # Spearman pooled
        rho, p_sp = stats.spearmanr(level_nums, norm)
        print(f"  Spearman (pooled n=20): ρ={rho:+.3f}, p={p_sp:.4f}  {sig(p_sp)}")

        # Kruskal-Wallis
        groups_l = [[d[('norm_mv_into' if 'into' in label
                          else 'norm_mv_outof' if 'out' in label
                          else 'norm_pv')]
                      for d in rows if d['lang'] == lang]
                     for lang in LANGUAGES]
        H, p_kw = stats.kruskal(*groups_l)
        print(f"  Kruskal-Wallis (L1): H(3)={H:.3f}, p={p_kw:.4f}  {sig(p_kw)}")

        # Per-language Spearman
        print("  Spearman per language (n=5):")
        for lang in LANGUAGES:
            lr = [d for d in rows if d['lang'] == lang]
            col = ('norm_mv_into' if 'into' in label
                   else 'norm_mv_outof' if 'out' in label
                   else 'norm_pv')
            rho_l, p_l = stats.spearmanr(
                [d['level_num'] for d in lr], [d[col] for d in lr])
            print(f"    {lang:<12}: ρ={rho_l:+.3f}, p={p_l:.4f}  {sig(p_l)}")

    # ── MV vs PV comparison ───────────────────────────────────────────────────
    print(f"\n{'='*65}")
    print("  MV constructions vs Path verbs — comparison")
    print(f"{'='*65}")
    rho_c, p_c = stats.spearmanr(norm_mv, norm_pv)
    print(f"\n  Spearman MV~PV (raw): ρ={rho_c:+.3f}, p={p_c:.4f}  {sig(p_c)}")
    rho_p, p_p = partial_spearman(norm_mv, norm_pv, level_nums)
    print(f"  Partial Spearman (ctrl. level): ρ={rho_p:+.3f}, p={p_p:.4f}  {sig(p_p)}")

    # Delta analysis
    transitions = [('A1','A2'),('A2','B1'),('B1','B2'),('B2','C1')]
    dm = []; dp = []
    for lang in LANGUAGES:
        for l1, l2 in transitions:
            d1 = results[(lang, l1)]; d2 = results[(lang, l2)]
            dm.append(d2['norm_mv_total'] - d1['norm_mv_total'])
            dp.append(d2['norm_pv']       - d1['norm_pv'])
    rho_d, p_d = stats.spearmanr(dm, dp)
    print(f"  Delta Spearman ΔMV vs ΔPV (n=16): ρ={rho_d:+.3f}, "
          f"p={p_d:.4f}  {sig(p_d)}")

    print("\n  Satellite-framing share by level:")
    print(f"  {'Level':<6} {'PV':>8} {'MV_into':>9} {'MV_outof':>10} "
          f"{'MV_tot':>8} {'MV/(MV+PV)':>12}")
    for lvl in LEVELS:
        pv_m  = np.mean([d['norm_pv']       for d in rows if d['level'] == lvl])
        mvi_m = np.mean([d['norm_mv_into']   for d in rows if d['level'] == lvl])
        mvo_m = np.mean([d['norm_mv_outof']  for d in rows if d['level'] == lvl])
        mvt_m = mvi_m + mvo_m
        rat   = mvt_m / (mvt_m + pv_m) * 100 if (mvt_m + pv_m) > 0 else 0
        print(f"  {lvl:<6} {pv_m:>8.4f} {mvi_m:>9.4f} {mvo_m:>10.4f} "
              f"{mvt_m:>8.4f} {rat:>11.1f}%")


# ╔══════════════════════════════════════════════════════════════════════════════
# 7. CONCORDANCE OUTPUT
# ╚══════════════════════════════════════════════════════════════════════════════

def save_concordances(results):
    """Save concordance lines for all three constructions for manual review."""
    for key_name, hit_key, fname in [
        ('MV+into',  'mv_into_hits',  'concordance_mv_into.txt'),
        ('MV+outof', 'mv_outof_hits', 'concordance_mv_outof.txt'),
        ('PV',       'pv_hits',       'concordance_pv.txt'),
    ]:
        with open(fname, 'w', encoding='utf-8') as f:
            f.write(f"Concordance: {key_name}\n{'='*80}\n\n")
            for lang in LANGUAGES:
                for level in LEVELS:
                    key = (lang, level)
                    if key not in results:
                        continue
                    hits = results[key][hit_key]
                    if hits:
                        f.write(f"--- {lang.upper()} {level} "
                                f"({len(hits)} hits) ---\n")
                        for item in hits:
                            lemma, ground, ctx = item[0], item[1], item[2]
                            clean = ' '.join(t.rsplit('_',1)[0]
                                             for t in ctx.split())
                            f.write(f"  [{lemma} | {ground}]  {clean}\n")
                        f.write('\n')
    print("Concordance files saved.")


# ╔══════════════════════════════════════════════════════════════════════════════
# 8. FIGURES
# ╚══════════════════════════════════════════════════════════════════════════════

def make_figures(results):
    """Generate and save all 5 publication figures."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    colors  = ['#2F5F9E', '#E07B39', '#3A9E5F', '#C0392B']
    markers = ['o', 's', '^', 'D']
    lang_labels = [l.capitalize() for l in LANGUAGES]
    x = np.arange(len(LEVELS))

    def get_norm(lang, metric):
        return [results[(lang, lvl)][metric] for lvl in LEVELS]

    def level_means(metric):
        return [np.mean([results[(l, lvl)][metric] for l in LANGUAGES])
                for lvl in LEVELS]

    # ── Figure 1: MV+into ────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(8, 5))
    for lang, col, mk, lbl in zip(LANGUAGES, colors, markers, lang_labels):
        ax.plot(x, get_norm(lang, 'norm_mv_into'), color=col, marker=mk,
                linewidth=2.2, markersize=8, label=lbl)
    ax.set_xticks(x); ax.set_xticklabels(LEVELS, fontsize=12)
    ax.set_xlabel('CEFR Proficiency Level', fontsize=12)
    ax.set_ylabel('Frequency per 10,000 words', fontsize=12)
    ax.set_title('Figure 1. Manner Verb + INTO (verified spatial Grounds)\n'
                 'across CEFR Levels by L1 Background',
                 fontsize=11, fontweight='bold')
    ax.legend(title='L1', fontsize=10, title_fontsize=10)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.set_ylim(-0.03, 0.65)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    plt.tight_layout()
    plt.savefig('Figure1_MV_into.png', dpi=300, bbox_inches='tight')
    plt.close()

    # ── Figure 2: MV+out of ──────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(8, 5))
    for lang, col, mk, lbl in zip(LANGUAGES, colors, markers, lang_labels):
        ax.plot(x, get_norm(lang, 'norm_mv_outof'), color=col, marker=mk,
                linewidth=2.2, markersize=8, label=lbl)
    ax.set_xticks(x); ax.set_xticklabels(LEVELS, fontsize=12)
    ax.set_xlabel('CEFR Proficiency Level', fontsize=12)
    ax.set_ylabel('Frequency per 10,000 words', fontsize=12)
    ax.set_title('Figure 2. Manner Verb + OUT OF (verified spatial Grounds)\n'
                 'across CEFR Levels by L1 Background',
                 fontsize=11, fontweight='bold')
    ax.legend(title='L1', fontsize=10, title_fontsize=10)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.set_ylim(-0.01, 0.22)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    plt.tight_layout()
    plt.savefig('Figure2_MV_outof.png', dpi=300, bbox_inches='tight')
    plt.close()

    # ── Figure 3: Path verbs ─────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(8, 5))
    for lang, col, mk, lbl in zip(LANGUAGES, colors, markers, lang_labels):
        ax.plot(x, get_norm(lang, 'norm_pv'), color=col, marker=mk,
                linewidth=2.2, markersize=8, label=lbl)
    ax.set_xticks(x); ax.set_xticklabels(LEVELS, fontsize=12)
    ax.set_xlabel('CEFR Proficiency Level', fontsize=12)
    ax.set_ylabel('Frequency per 10,000 words', fontsize=12)
    ax.set_title('Figure 3. Boundary-Crossing Path Verbs (enter + exit,\n'
                 'verified spatial Grounds) across CEFR Levels by L1',
                 fontsize=11, fontweight='bold')
    ax.legend(title='L1', fontsize=10, title_fontsize=10)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.set_ylim(-0.02, 0.65)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    plt.tight_layout()
    plt.savefig('Figure3_PathVerbs.png', dpi=300, bbox_inches='tight')
    plt.close()

    # ── Figure 4: Comparison ─────────────────────────────────────────────────
    pv_m   = level_means('norm_pv')
    mvi_m  = level_means('norm_mv_into')
    mvo_m  = level_means('norm_mv_outof')
    mvt_m  = [mvi_m[i] + mvo_m[i] for i in range(5)]
    ratio  = [mvt_m[i] / (mvt_m[i] + pv_m[i]) * 100
              if (mvt_m[i] + pv_m[i]) > 0 else 0 for i in range(5)]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))
    ax1.plot(x, pv_m,  'o-',  color='#C0392B', linewidth=2.5, markersize=9,
             label='Path verbs (enter + exit)')
    ax1.plot(x, mvi_m, 's--', color='#2F5F9E', linewidth=2.2, markersize=8,
             label='MV + into')
    ax1.plot(x, mvo_m, '^:',  color='#3A9E5F', linewidth=2.0, markersize=8,
             label='MV + out of')
    ax1.set_xticks(x); ax1.set_xticklabels(LEVELS, fontsize=11)
    ax1.set_xlabel('CEFR Level', fontsize=11)
    ax1.set_ylabel('Mean freq. per 10,000 words', fontsize=11)
    ax1.set_title('Absolute Rates\n(mean across 4 L1 groups)',
                  fontsize=11, fontweight='bold')
    ax1.legend(fontsize=9)
    ax1.grid(True, alpha=0.3, linestyle='--')
    ax1.spines['top'].set_visible(False); ax1.spines['right'].set_visible(False)

    ax2.bar(x, ratio, color='#2F5F9E', alpha=0.75,
            edgecolor='white', linewidth=0.5)
    ax2.set_xticks(x); ax2.set_xticklabels(LEVELS, fontsize=11)
    ax2.set_xlabel('CEFR Level', fontsize=11)
    ax2.set_ylabel('MV constructions as % of\ntotal boundary-crossing expressions',
                   fontsize=10)
    ax2.set_title('Satellite-Framing Share of\nBoundary-Crossing Expressions',
                  fontsize=11, fontweight='bold')
    ax2.set_ylim(0, 85)
    for i, v in enumerate(ratio):
        ax2.text(i, v + 1.5, f'{v:.1f}%', ha='center', fontsize=11,
                 fontweight='bold', color='#1a1a1a')
    ax2.grid(True, alpha=0.3, linestyle='--', axis='y')
    ax2.spines['top'].set_visible(False); ax2.spines['right'].set_visible(False)
    fig.suptitle('Figure 4. Manner Verb Constructions vs. Path Verbs '
                 'across Proficiency Levels',
                 fontsize=11, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig('Figure4_Comparison.png', dpi=300, bbox_inches='tight')
    plt.close()

    # ── Figure 5: SM vs CM ───────────────────────────────────────────────────
    sm_into = {
        lang: [
            sum(1 for lemma, _, _ in results[(lang, lvl)]['mv_into_hits']
                if get_motion_type(lemma) == 'self-motion')
            for lvl in LEVELS
        ] for lang in LANGUAGES
    }
    cm_into = {
        lang: [
            sum(1 for lemma, _, _ in results[(lang, lvl)]['mv_into_hits']
                if get_motion_type(lemma) == 'caused-motion')
            for lvl in LEVELS
        ] for lang in LANGUAGES
    }

    fig, axes = plt.subplots(1, 4, figsize=(14, 4.5), sharey=False)
    for lang, ax, lbl in zip(LANGUAGES, axes, lang_labels):
        ax.bar(x - 0.175, sm_into[lang], 0.35, label='Self-motion',
               color='#5B9BD5', edgecolor='white', linewidth=0.5)
        ax.bar(x + 0.175, cm_into[lang], 0.35, label='Caused-motion',
               color='#F4A261', edgecolor='white', linewidth=0.5)
        ax.set_title(lbl, fontsize=11, fontweight='bold')
        ax.set_xticks(x); ax.set_xticklabels(LEVELS, fontsize=9)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.25, axis='y', linestyle='--')
        if lang == LANGUAGES[0]:
            ax.set_ylabel('Raw count (verified)', fontsize=10)

    handles = [mpatches.Patch(color='#5B9BD5', label='Self-motion'),
               mpatches.Patch(color='#F4A261', label='Caused-motion')]
    fig.legend(handles=handles, loc='lower center', ncol=2, fontsize=10,
               title='Motion type', title_fontsize=10,
               bbox_to_anchor=(0.5, -0.05))
    fig.suptitle('Figure 5. Self-motion vs. Caused-motion Verbs in MV+into\n'
                 '(verified, by L1 and CEFR level)',
                 fontsize=11, fontweight='bold', y=1.03)
    plt.tight_layout()
    plt.savefig('Figure5_SM_CM.png', dpi=300, bbox_inches='tight')
    plt.close()

    print("All 5 figures saved (300 dpi PNG).")


# ╔══════════════════════════════════════════════════════════════════════════════
# 9. EXCEL OUTPUT
# ╚══════════════════════════════════════════════════════════════════════════════

def save_excel(results, fname='manner_verb_results.xlsx'):
    """Save frequency tables to Excel with colour-coded columns."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'

    hfill = PatternFill('solid', fgColor='1F3864')
    hfont = Font(bold=True, color='FFFFFF', size=10)
    afill = PatternFill('solid', fgColor='EEF2FF')
    bl_fill = PatternFill('solid', fgColor='D6E4F0')
    or_fill = PatternFill('solid', fgColor='FCE8D5')
    gn_fill = PatternFill('solid', fgColor='D5EAD5')

    headers = ['Language', 'Level', 'Word Count',
               'Raw MV+into', 'Norm MV+into (/10k)',
               'Raw MV+outof', 'Norm MV+outof (/10k)',
               'Raw PV (enter+exit)', 'Norm PV (/10k)',
               'SM MV+into', 'CM MV+into']

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = (bl_fill if 'into' in h else
                     or_fill if 'outof' in h else
                     gn_fill if 'PV' in h else hfill)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    ri = 2
    for lang in LANGUAGES:
        for level in LEVELS:
            key = (lang, level)
            if key not in results:
                continue
            d = results[key]
            sm = sum(1 for lemma, _, _ in d['mv_into_hits']
                     if get_motion_type(lemma) == 'self-motion')
            cm = sum(1 for lemma, _, _ in d['mv_into_hits']
                     if get_motion_type(lemma) == 'caused-motion')
            row_vals = [
                lang.capitalize(), level, d['wc'],
                d['raw_mv_into'],  d['norm_mv_into'],
                d['raw_mv_outof'], d['norm_mv_outof'],
                d['raw_pv'],       d['norm_pv'],
                sm, cm,
            ]
            fill = afill if ri % 2 == 0 else None
            for ci, v in enumerate(row_vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.alignment = Alignment(horizontal='center')
                if fill:
                    cell.fill = fill
            ri += 1

    col_widths = [14, 8, 14, 14, 20, 14, 22, 20, 14, 14, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(fname)
    print(f"Excel saved: {fname}")


# ╔══════════════════════════════════════════════════════════════════════════════
# 10. MAIN
# ╚══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print('=' * 65)
    print('BOUNDARY-CROSSING MOTION EVENT ANALYSIS — L2 ENGLISH')
    print('=' * 65)

    print('\n[1/5] Running corpus search...')
    results = run_corpus_search()

    print('\n[2/5] Saving concordance files...')
    save_concordances(results)

    print('\n[3/5] Running statistical analyses...')
    run_statistics(results)

    print('\n[4/5] Generating figures...')
    make_figures(results)

    print('\n[5/5] Saving Excel output...')
    save_excel(results)

    print('\nAll done.')
